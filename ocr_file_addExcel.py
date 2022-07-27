from pytesseract import Output      #reading an image 
import pytesseract
import cv2
import shutil                      #working with directory
import os                        
from PIL import Image
import os.path, sys
from time import sleep
from tqdm import tqdm               #showing progress bar in cmd
from openpyxl import Workbook       #creat and work with excel workbook
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder


# drawings for example saved in the folder dwg
folder_path = "C:\\Users\\***\\***\\***\\dwg"      #THIS SHOULD BE CHANGED DEPENDS OF ORIGINAL/ DST FOLDER LOCATION
dirs = os.listdir(folder_path)

total_count = len([name for name in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, name))])  #total number of the files in the directory


wb = Workbook()                            #creating new excel workbook to store summary 
ws = wb.active
ws.title = "Drawings Summary"              #name and properties of the worksheet
ws.sheet_properties.tabColor = "1072BA"

ws.merge_cells('A1:E1')
cell_title = ws.cell(row=1, column=1)
cell_title.value = 'List of Drawings'
cell_title.alignment = Alignment(horizontal='center', vertical='center')
cell_title.font = Font(bold=True, size=16, color='00333399')

header_list = ['plumbing_dwg', 'elev_dwg', 'cellar_dwg', 'roof_dwg', 'borings_dwg']
col = 1
for header in list(header_list):
    ws.cell(row = 2, column = col).value = str(header).capitalize()
    col += 1   
h_font = Font(bold=True, italic=True, size=12)
for cell_h in ws['2:2']:
    cell_h.font = h_font

dim_holder = DimensionHolder(ws)
for col in range(ws.min_column, ws.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
ws.column_dimensions = dim_holder


plumbing_dwg = 0          #count how many drawings in each folder after files moved
elev_dwg = 0
cellar_dwg = 0
roof_dwg = 0
borings_dwg = 0
other_dwg = 0

def crop():
    try:
        for item in tqdm(dirs):
            sleep(0.05)
            global dwg_path
            dwg_path = os.path.join(folder_path,item)                   #drawing file path
                        
            if os.path.isfile(dwg_path):                
            
                im = Image.open(dwg_path)
                
                width, height = im.size                      
                left = width - (width/6)
                right = width
                top = height - (height/4)
                bottom = height
                cropped_im = im.crop((left, top, right, bottom))        #croping bottom right corner of the drawing
                #cropped_im.show()
                
                myconfig = r"-l eng --psm 6 --oem 3"
                text = pytesseract.image_to_string(cropped_im, config=myconfig) #eng - using english language; psm 6 - Page segmentation mode - Assume a single uniform block of text; oem 3 - OCR Engine Mode - Legacy + LSTM engines

                def moving_file():
                    global plumbing_dwg
                    global elev_dwg
                    global cellar_dwg
                    global roof_dwg
                    global borings_dwg
                    global other_dwg
                    
                    t = text.lower()
                    if 'plumbing' in t or 'drainage' in t:              #searching for specific words in text
                        dst_path = os.path.join(folder_path,'plumbing') #creating new folder path 
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                       #creating new folder (if doesn't have one)
                        shutil.move(dwg_path, dst_path)                 #moving file into the folder
                        plumbing_dwg += 1
                        cl_A = max((a.row for a in ws['A'] if a.value is not None))
                        ws.cell(column=1, row=cl_A+1, value=f'{item}')  #saving file name in rep column                       
                    elif 'elevation' in t or 'survey' in t:                                
                        dst_path = os.path.join(folder_path,'elevation') 
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                                         
                        shutil.move(dwg_path, dst_path)
                        elev_dwg += 1
                        cl_B = max((b.row for b in ws['B'] if b.value is not None)) 
                        ws.cell(column=2, row=cl_B+1, value=f'{item}')
                    elif 'cellar' in t or 'basement' in t:                                
                        dst_path = os.path.join(folder_path,'cellar') 
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                                         
                        shutil.move(dwg_path, dst_path) 
                        cellar_dwg += 1                        
                        cl_C = max((c.row for c in ws['C'] if c.value is not None))
                        ws.cell(column=3, row=cl_C+1, value=f'{item}')                         
                    elif 'roof' in t:                                
                        dst_path = os.path.join(folder_path,'roof')
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                                         
                        shutil.move(dwg_path, dst_path)
                        roof_dwg += 1
                        cl_D = max((d.row for d in ws['D'] if d.value is not None))
                        ws.cell(column=4, row=cl_D+1, value=f'{item}') 
                    elif 'borings' in t:                                
                        dst_path = os.path.join(folder_path,'borings')
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                                         
                        shutil.move(dwg_path, dst_path)
                        borings_dwg += 1
                        cl_E = max((e.row for e in ws['E'] if e.value is not None))
                        ws.cell(column=5, row=cl_E+1, value=f'{item}') 
                    else:                               
                        dst_path = os.path.join(folder_path,'other_dwg')
                        if not os.path.exists(dst_path):
                            os.makedirs(dst_path)                                         
                        shutil.move(dwg_path, dst_path)
                        other_dwg += 1
        
                moving_file()               
    except IOError:           #raise an IOError if file cannot be found, or image cannot be opened.
        pass


print('Total number of drawings in the folder: ', total_count)
crop()
print('Plumbing drawings: ', plumbing_dwg)
print('Elevation drawings: ', elev_dwg)
print('Cellar drawings: ', cellar_dwg)
print('Roof drawings: ', roof_dwg)
print('Borings drawings: ', borings_dwg)
print('Other drawings: ', other_dwg)

wb.save('dwg\\Summary.xlsx')
wb.close()


